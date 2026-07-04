import base64
import os
from datetime import datetime

import anthropic
import openpyxl
from PIL import Image
from pillow_heif import register_heif_opener

register_heif_opener()

client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))

with open("job_list.txt", "r") as job_list_file:
    job_list_contents = job_list_file.read()


def convert_image(image_path):
    if image_path.endswith(".heic") or image_path.endswith(".HEIC"):
        img = Image.open(image_path)
        new_path = image_path.replace(".heic", ".jpg").replace(".HEIC", ".jpg")
        img.save(new_path, "JPEG")
        os.remove(image_path)
        return new_path
    return image_path


def read_image(image_path):
    image_path = convert_image(image_path)
    with open(image_path, "rb") as image_file:
        image_data = image_file.read()
    return image_data, image_path


def parse_truck(value):
    """Convert truck label to a whole number."""
    return int(float(value.strip()))


def parse_hours(value):
    """Convert hours text to a number, allowing decimals like 8.5."""
    return float(value.strip())


def process_timesheet(image_path):
    image_data, image_path = read_image(image_path)

    image_base64 = base64.b64encode(image_data).decode("utf-8")

    message = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=1024,
        messages=[
            {"role": "user", "content": [
                {
                    "type": "image",
                    "source": {
                        "type": "base64",
                        "media_type": "image/jpeg",
                        "data": image_base64
                    }
                },
                {
                    "type": "text",
                    "text": """You are a helpful assistant for a landscaping business.

Here is the official job list:
""" + job_list_contents + """

Extract the timesheet data from the image and return it in this exact format:
EMPLOYEE: [name]
DATE: [date in MM/DD/YYYY format]
JOB: [official job name] | TRUCK: [truck number only, just the number] | HOURS: [hours as a number, use decimals when needed e.g. 8.0 or 8.5]"""
                }
            ]}
        ]
    )
    return message.content[0].text


def update_spreadsheet(claude_response):
    try:
        workbook = openpyxl.load_workbook("timesheets_final.xlsx")
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet["A1"] = "Employee"
        sheet["B1"] = "Date"
        sheet["C1"] = "Job Site"
        sheet["D1"] = "Truck"
        sheet["E1"] = "Hours"
        sheet["F1"] = "Date Submitted"

    lines = claude_response.split("\n")
    employee = ""
    date = ""
    row = sheet.max_row + 1
    rows_added = 0

    for line in lines:
        if line.startswith("EMPLOYEE:"):
            employee = line.replace("EMPLOYEE: ", "").strip()
        elif line.startswith("DATE:"):
            date = line.replace("DATE: ", "").strip()
        elif line.startswith("JOB:"):
            parts = line.replace("JOB: ", "").split(" | ")
            if len(parts) < 3:
                raise ValueError(f"Could not parse job line (expected 3 parts): {line}")

            job = parts[0].strip()
            truck = parts[1].replace("TRUCK: ", "").strip()
            hours = parts[2].replace("HOURS: ", "").strip()

            sheet["A" + str(row)] = employee
            sheet["B" + str(row)] = date
            sheet["C" + str(row)] = job
            sheet["D" + str(row)] = parse_truck(truck)
            sheet["E" + str(row)] = parse_hours(hours)
            sheet["F" + str(row)] = datetime.now().strftime("%m/%d/%Y %I:%M %p")
            row = row + 1
            rows_added = rows_added + 1

    if rows_added == 0:
        raise ValueError("No job rows found in Claude's response. Check the image or try again.")

    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 15
    sheet.column_dimensions["C"].width = 35
    sheet.column_dimensions["D"].width = 10
    sheet.column_dimensions["E"].width = 15
    sheet.column_dimensions["F"].width = 25

    workbook.save("timesheets_final.xlsx")
    print(f"Spreadsheet updated! Added {rows_added} row(s).")
    return rows_added
